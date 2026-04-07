"""
URL Kategorisierung - Spalte E (Art des Inhalts genau)
Liest Excel, ruft jede URL auf, bestimmt Kategorie, speichert Ergebnis.
"""

import os
import time
import requests
import anthropic
import openpyxl
from pathlib import Path

# ─── Konfiguration ────────────────────────────────────────────────────────────
EXCEL_DATEI   = "AI Coaching Tools and Knowledge.xlsx"   # Dateiname anpassen
TABELLENBLATT = "aktuell"
SPALTE_URL    = 1   # Spalte A
SPALTE_E      = 5   # Spalte E = Art des Inhalts genau
HEADER_ZEILEN = 2   # Zeile 1 = Überschriften, Zeile 2 = Prompts → nicht anfassen
PAUSE_SEKUNDEN = 1  # Pause zwischen API-Aufrufen (verhindert Rate Limiting)

# Anthropic API Key aus Umgebungsvariable oder direkt eintragen
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

KATEGORIEN = ["Produktseite", "Artikel", "Studie", "sonstiges"]

# ─── Hilfsfunktionen ──────────────────────────────────────────────────────────
def hole_seiteninhalt(url: str) -> str:
    """Ruft eine URL auf und gibt den Text-Inhalt zurück."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; KI-Coaching-Kompass/1.0)"}
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        # Nur die ersten 3000 Zeichen — reicht für Kategorisierung
        return response.text[:3000]
    except Exception as e:
        return f"FEHLER: {e}"


def kategorisiere(url: str, seiteninhalt: str, client: anthropic.Anthropic) -> str:
    """Fragt Claude welche Kategorie diese URL hat."""
    if seiteninhalt.startswith("FEHLER"):
        return "nicht erreichbar"

    prompt = f"""Analysiere den folgenden Webseiteninhalt und bestimme die Art des Inhalts.
Verwende AUSSCHLIESSLICH einen dieser Begriffe als Antwort:
- Produktseite (wenn es sich um ein Tool, Software oder Produkt handelt)
- Artikel (wenn es sich um einen Blog-Post, News-Artikel oder Meinungsbeitrag handelt)
- Studie (wenn es sich um eine wissenschaftliche Publikation, Whitepaper oder Forschungsbericht handelt)
- sonstiges (alles andere)

Antworte NUR mit dem Begriff, ohne Erklärung, ohne Satzzeichen.

URL: {url}

Webseiteninhalt (Auszug):
{seiteninhalt[:2000]}"""

    try:
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=20,
            messages=[{"role": "user", "content": prompt}]
        )
        antwort = message.content[0].text.strip()
        # Sicherstellen dass nur gültige Kategorien zurückkommen
        for kategorie in KATEGORIEN:
            if kategorie.lower() in antwort.lower():
                return kategorie
        return "sonstiges"
    except Exception as e:
        return f"API-Fehler: {e}"


# ─── Hauptprogramm ─────────────────────────────────────────────────────────────
def main():
    if not API_KEY:
        print("❌ ANTHROPIC_API_KEY nicht gesetzt!")
        print("   Setze ihn mit: set ANTHROPIC_API_KEY=sk-ant-...")
        return

    # Excel-Datei finden
    excel_pfad = Path(EXCEL_DATEI)
    if not excel_pfad.exists():
        print(f"❌ Datei nicht gefunden: {EXCEL_DATEI}")
        print(f"   Aktuelles Verzeichnis: {Path.cwd()}")
        return

    print(f"📂 Öffne: {EXCEL_DATEI}")
    wb = openpyxl.load_workbook(excel_pfad)

    if TABELLENBLATT not in wb.sheetnames:
        print(f"❌ Tabellenblatt '{TABELLENBLATT}' nicht gefunden.")
        print(f"   Verfügbare Blätter: {wb.sheetnames}")
        return

    ws = wb[TABELLENBLATT]
    client = anthropic.Anthropic(api_key=API_KEY)

    # Zeilen zählen
    gesamt = sum(1 for row in ws.iter_rows(min_row=HEADER_ZEILEN + 1, values_only=True)
                 if row[SPALTE_URL - 1])
    print(f"📊 {gesamt} URLs gefunden. Starte Kategorisierung...\n")

    verarbeitet = 0
    uebersprungen = 0

    for zeile_nr, row in enumerate(ws.iter_rows(min_row=HEADER_ZEILEN + 1), start=HEADER_ZEILEN + 1):
        url_zelle = row[SPALTE_URL - 1]
        e_zelle   = row[SPALTE_E - 1]

        url = url_zelle.value if url_zelle else None

        # Leere URLs überspringen
        if not url or not str(url).startswith("http"):
            continue

        # Bereits befüllte Zeilen überspringen
        if e_zelle and e_zelle.value and e_zelle.value not in ["", None]:
            uebersprungen += 1
            continue

        verarbeitet += 1
        print(f"[{verarbeitet}/{gesamt}] Zeile {zeile_nr}: {url[:60]}...")

        # Seiteninhalt holen
        inhalt = hole_seiteninhalt(url)

        # Kategorisieren
        kategorie = kategorisiere(url, inhalt, client)
        print(f"         → {kategorie}")

        # In Excel schreiben
        ws.cell(row=zeile_nr, column=SPALTE_E).value = kategorie

        # Nach jeder 10. Zeile speichern
        if verarbeitet % 10 == 0:
            wb.save(excel_pfad)
            print(f"         💾 Gespeichert ({verarbeitet} Zeilen verarbeitet)")

        time.sleep(PAUSE_SEKUNDEN)

    # Abschließend speichern
    wb.save(excel_pfad)
    print(f"\n✅ Fertig! {verarbeitet} URLs kategorisiert, {uebersprungen} bereits befüllt übersprungen.")
    print(f"📁 Gespeichert: {excel_pfad.resolve()}")


if __name__ == "__main__":
    main()
